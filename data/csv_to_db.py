import sqlite3
import pandas as pd
import logging
from datetime import datetime
from openpyxl import load_workbook

with open('data/data.db','wb+') as file_obj:
    file_obj.close()
con = sqlite3.connect('data/data.db')

logger = logging.getLogger(__name__)


file_handler = logging.FileHandler(f"logs/CONVERT_{datetime.now().strftime('%Y-%m-%d-%H-%M')}.log")
file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)-5.5s]  %(message)s"))

logger.addHandler(file_handler)
logger.addHandler(logging.StreamHandler())
logger.setLevel(logging.DEBUG)

def update_investment():
    df_invest = pd.read_csv('data/inve.csv')
    df_invest['date'] = pd.to_datetime(df_invest['date']).dt.date
    df_invest.to_sql('investment',con, if_exists='append', index=False)
    logger.info('Investment data successfully stored in DB!')

def xlsx_to_csv():
    wb = load_workbook(filename='data/budget.xlsx')
    sheets_list = wb.sheetnames
    all_sheets = []
    for name in sheets_list[::-1]:
        df = pd.DataFrame(wb[name].values)
        df = df.iloc[:,:3]
        df = df.rename(columns={1:'date',0:'item',2:'price'})
        df = df.dropna()
        df['date'] = pd.to_datetime(df['date']).dt.date
        try:
            df['price'] = df['price'].map(lambda x: x.strip('='))
        except AttributeError:
            print(f'Check sheet:{name}')
        df['price'] = df['price'].str.split('+')
        df['item'] = df['item'].str.split(',')
        print(f'Sheet name: {name}')
        df = df.explode(['price','item'])
        all_sheets.append(df)
        if name == 'March18':
            break
    df_cons = pd.concat(all_sheets, ignore_index=True)
    df_cons = df_cons.sort_values(by='date')
    print(df_cons)
    df_cons.info()
    df_cons.to_csv('df_cons.csv', index=False)



if __name__ == '__main__':
    xlsx_to_csv()